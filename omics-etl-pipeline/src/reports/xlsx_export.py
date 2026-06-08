"""
Exportação XLSX consolidada — Omics ETL Pipeline.

Gera um workbook com 6 abas sem novas consultas ao banco:
  1. Resumo          — status, risco, diagnóstico (todos os públicos)
  2. Resultado_Final — um Rank 1 por composto com scores (todos)
  3. Para_Revisao    — compostos Alta/Média com ação recomendada (especialistas)
  4. Dados_Tecnicos  — todos os candidatos e ranks (especialistas)
  5. Estatisticas    — distribuições agregadas do experimento (todos)
  6. Metadados       — rastreabilidade e auditoria (todos)

Entrada: ins (computar_insights), nar (gerar_narrativa), batch_info, cobertura_ext
Saída:   bytes de workbook XLSX via BytesIO (geração 100% em memória)
"""
from __future__ import annotations

import re
from datetime import datetime
from io import BytesIO
from typing import TYPE_CHECKING

import pandas as pd

if TYPE_CHECKING:
    from src.reports.narrative import NarrativaDict

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    _HAS_OPENPYXL = True
except ImportError:
    _HAS_OPENPYXL = False

# ─── Paleta de cores ────────────────────────────────────────────────────────
_C = {
    "hdr_bg":   "003366",   # Azul IST — cabeçalho de tabela
    "hdr_fg":   "FFFFFF",
    "sect_bg":  "1f4e79",   # Azul escuro — faixa de seção
    "sect_fg":  "FFFFFF",
    "sub_bg":   "e8e8e8",   # Cinza — sub-seção interna
    "sub_fg":   "3d3d3d",
    "alt":      "f2f2f2",   # Cinza alternado
    "white":    "FFFFFF",
    "alta":     "f8d7da",   # Vermelho claro — Alta prioridade
    "media":    "fff3cd",   # Amarelo claro — Média prioridade
    "ok":       "d4edda",   # Verde claro — OK / conclusivo
    "conclusivo":   "d4edda",
    "parcial":      "fff3cd",
    "inconclusivo": "f8d7da",
    "Baixo":    "d4edda",
    "Moderado": "fff3cd",
    "Elevado":  "f8d7da",
}

# ─── Mapeamentos de prioridade ────────────────────────────────────────────
_PRIO_LABEL = {"🔴 Alta": "Alta", "🟡 Média": "Média", "✅ OK": "OK", "—": "—"}
_PRIO_COLOR = {"Alta": "alta", "Média": "media", "OK": "ok", "—": "white"}
_PRIO_SORT  = {"Alta": 0, "Média": 1, "OK": 2, "—": 3}

_STATUS_LABEL = {
    "conclusivo":   "ANÁLISE CONCLUSIVA",
    "parcial":      "ANÁLISE PARCIALMENTE CONCLUSIVA",
    "inconclusivo": "ANÁLISE INCONCLUSIVA",
}

_ACAO = {
    "Alta":  "Comparar espectros manualmente no software do instrumento e selecionar o candidato correto.",
    "Média": "Avaliar espectro individual e confirmar se a identificação automática está correta.",
}

_STRIP_MD = re.compile(r"\*{1,2}([^*]+)\*{1,2}")


def _md(text: str) -> str:
    """Remove markdown bold/italic for plain cell values."""
    return _STRIP_MD.sub(r"\1", str(text)) if text else ""


# ─── Helpers de estilo ────────────────────────────────────────────────────

def _fill(hex_color: str) -> "PatternFill":
    return PatternFill(fgColor=hex_color, fill_type="solid")


def _font(bold: bool = False, color: str = "000000",
          size: int = 10, name: str = "Calibri") -> "Font":
    return Font(bold=bold, color=color, size=size, name=name)


def _align(h: str = "left", v: str = "center", wrap: bool = False) -> "Alignment":
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def _norm(val):
    """Normaliza valor numpy/NaN para tipo Python nativo."""
    if val is None:
        return None
    if hasattr(val, "item"):
        val = val.item()
    try:
        if pd.isna(val):
            return None
    except (TypeError, ValueError):
        pass
    return val


def _write_header(ws, cols: list[str], row: int = 1) -> None:
    for c, name in enumerate(cols, 1):
        cell = ws.cell(row=row, column=c, value=name)
        cell.fill  = _fill(_C["hdr_bg"])
        cell.font  = _font(bold=True, color=_C["hdr_fg"])
        cell.alignment = _align(h="center")


def _sect(ws, row: int, title: str, ncols: int = 2, bg: str = "sect_bg") -> int:
    """Escreve faixa de seção e retorna próxima linha."""
    cell = ws.cell(row=row, column=1, value=f"  {title}")
    cell.fill = _fill(_C[bg])
    cell.font = _font(bold=True, color=_C["sect_fg"] if bg == "sect_bg" else _C["sub_fg"], size=11)
    cell.alignment = _align(h="left")
    if ncols > 1:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    return row + 1


def _kv(ws, row: int, key: str, val, val_fill: str | None = None) -> int:
    """Escreve par chave-valor e retorna próxima linha."""
    ck = ws.cell(row=row, column=1, value=key)
    cv = ws.cell(row=row, column=2, value=_norm(val))
    ck.font = _font(bold=True)
    ck.alignment = _align(h="right")
    cv.alignment = _align()
    if val_fill:
        cv.fill = _fill(val_fill)
    return row + 1


def _auto_width(ws, min_w: int = 8, max_w: int = 52) -> None:
    """Ajusta largura das colunas com base no conteúdo (amostra header + 120 linhas)."""
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        sample = list(col)[:121]
        try:
            max_len = max((len(str(c.value or "")) for c in sample), default=min_w)
        except Exception:
            max_len = min_w
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, min_w), max_w)


def _write_table(ws, df: pd.DataFrame, start_row: int = 1,
                 priority_col: str | None = None) -> None:
    """Escreve cabeçalho + dados com cor por prioridade ou alternância cinza."""
    cols = list(df.columns)
    pc_idx = cols.index(priority_col) if priority_col in cols else None

    _write_header(ws, cols, start_row)

    for r_off, (_, series) in enumerate(df.iterrows()):
        r = start_row + 1 + r_off
        row_fill = None

        if pc_idx is not None:
            pval = str(series.iloc[pc_idx])
            row_fill = _C.get(_PRIO_COLOR.get(pval, "white"), _C["white"])
        elif r_off % 2 == 1:
            row_fill = _C["alt"]

        for c_off, val in enumerate(series):
            cell = ws.cell(row=r, column=c_off + 1, value=_norm(val))
            cell.alignment = _align()
            if row_fill:
                cell.fill = _fill(row_fill)

    # Auto-filtro após escrita
    if len(df) > 0:
        ws.auto_filter.ref = ws.dimensions


# ─── Aba 1: Resumo ────────────────────────────────────────────────────────

def _build_resumo(wb, ins: dict, nar: "NarrativaDict",
                  batch_info: dict, cobertura_ext: dict) -> None:
    ws = wb.create_sheet("Resumo")
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 55
    ws.column_dimensions["C"].width = 65

    n = 1

    # ── Título ──────────────────────────────────────────────────────────
    title = ws.cell(row=n, column=1,
                    value="Omics ETL Pipeline — Resultado do Experimento")
    title.font  = _font(bold=True, color=_C["hdr_fg"], size=14)
    title.fill  = _fill(_C["hdr_bg"])
    title.alignment = _align(h="left")
    ws.merge_cells(start_row=n, start_column=1, end_row=n, end_column=3)
    ws.row_dimensions[n].height = 26
    n += 2

    # ── Seção A: Identificação ───────────────────────────────────────────
    n = _sect(ws, n, "IDENTIFICAÇÃO DA ANÁLISE", ncols=3)
    n = _kv(ws, n, "Análise",
             f"#{batch_info.get('id', '—')}")
    d_ini = (batch_info.get("iniciado_em") or "")[:16].replace("T", " ") or "—"
    d_fim = (batch_info.get("concluido_em") or "")[:16].replace("T", " ") or "—"
    n = _kv(ws, n, "Processado em",  d_ini)
    n = _kv(ws, n, "Concluído em",   d_fim)
    n = _kv(ws, n, "Arquivo de identificação",
             batch_info.get("nome_ident", "—"))
    n = _kv(ws, n, "Arquivo de abundância",
             batch_info.get("nome_abund", "—"))
    n = _kv(ws, n, "Total de compostos",
             batch_info.get("total_sinais") or ins["n_compostos"])
    n = _kv(ws, n, "Total de candidatos avaliados",
             batch_info.get("total_candidatos") or ins["n_candidatos_tot"])
    n += 1

    # ── Seção B: Status e Risco ──────────────────────────────────────────
    n = _sect(ws, n, "STATUS E RISCO", ncols=3)
    status_lbl = _STATUS_LABEL.get(nar["status"], nar["status"].upper())
    n = _kv(ws, n, "Status da análise", status_lbl,
             val_fill=_C.get(nar["status"], _C["white"]))
    n = _kv(ws, n, "Risco analítico", nar["risco_label"],
             val_fill=_C.get(nar["risco_label"], _C["white"]))
    n = _kv(ws, n, "Descrição do risco", _md(nar["risco_desc"]))
    n = _kv(ws, n, "Ação recomendada",   _md(nar["recomendacao"]))
    n += 1

    # ── Seção C: Números-chave ───────────────────────────────────────────
    n = _sect(ws, n, "NÚMEROS-CHAVE", ncols=3)
    n_comp = ins["n_compostos"]
    n_rev  = nar["n_revisar"]
    n_alt  = ins["n_alta_conf"]
    n_emp  = ins["n_empates"]
    pct_rev = n_rev / n_comp * 100 if n_comp else 0
    pct_alt = n_alt / n_comp * 100 if n_comp else 0
    pct_emp = ins["pct_empates"]
    n = _kv(ws, n, "Compostos identificados", n_comp)
    n = _kv(ws, n, "Para revisão",
             f"{n_rev} de {n_comp}  ({pct_rev:.0f}%)")
    n = _kv(ws, n, "Alta confiança  (score ≥ 80)",
             f"{n_alt}  ({pct_alt:.0f}%)")
    n = _kv(ws, n, "Empates detectados",
             f"{n_emp}  ({pct_emp:.0f}%)")
    n += 1

    # ── Seção D: Diagnóstico Automático ──────────────────────────────────
    n = _sect(ws, n, "DIAGNÓSTICO AUTOMÁTICO", ncols=3)
    _write_header(ws, ["Dimensão", "Valor", "Interpretação"], row=n)
    n += 1
    for r_off, (dim, val, interp) in enumerate(nar["dimensoes"]):
        fill_color = _C["alt"] if r_off % 2 == 1 else _C["white"]
        for c_idx, txt in enumerate([dim, val, interp], 1):
            cell = ws.cell(row=n, column=c_idx, value=txt)
            cell.alignment = _align()
            cell.fill = _fill(fill_color)
            if c_idx == 1:
                cell.font = _font(bold=True)
        n += 1
    n += 1

    # ── Rodapé ───────────────────────────────────────────────────────────
    for c_idx, txt in enumerate(
        ["Exportado em",
         datetime.now().strftime("%Y-%m-%d %H:%M"),
         "Omics ETL Pipeline · IST Ambiental / SENAI"],
        1
    ):
        cell = ws.cell(row=n, column=c_idx, value=txt)
        cell.font = _font(color="888888", size=9)
        cell.alignment = _align()


# ─── Aba 2: Resultado_Final ──────────────────────────────────────────────

def _prep_resultado_df(ins: dict, nar: "NarrativaDict") -> pd.DataFrame:
    score_col = ins["score_col"]
    r1 = ins["rank1_unico"].copy()

    # Adiciona n_candidatos
    cd = ins["compound_data"][["Sinal", "n_candidatos"]].copy()
    r1 = r1.merge(cd, on="Sinal", how="left")

    # Adiciona Prioridade e Situação do priority_df
    pdf = nar["priority_df"][["Composto", "Prioridade", "Situação"]].copy()
    pdf.rename(columns={"Composto": "Sinal"}, inplace=True)
    r1 = r1.merge(pdf, on="Sinal", how="left")

    # Limpa rótulos
    r1["Prioridade"] = r1["Prioridade"].map(_PRIO_LABEL).fillna("—")
    if "Empate" in r1.columns:
        r1["Empate"] = r1["Empate"].map({1: "Sim", 0: "Não",
                                         True: "Sim", False: "Não"}).fillna("Não")

    # Ordena: prioridade → score desc
    r1["_ps"] = r1["Prioridade"].map(_PRIO_SORT).fillna(3)
    r1 = r1.sort_values(["_ps", score_col], ascending=[True, False]).drop("_ps", axis=1)

    desired = [
        ("Sinal",                "Composto"),
        ("m/z Medido",           "m/z Medido"),
        ("Candidato",            "Candidato Identificado"),
        ("Formula",              "Fórmula Molecular"),
        ("Peso Teorico",         "Peso Molecular (Da)"),
        ("Categoria",            "Categoria"),
        ("Classe Quimica",       "Classe Química (ChEBI)"),
        ("Prioridade",           "Prioridade"),
        ("Situação",             "Situação"),
        (score_col,              "Confiança (0–100)"),
        ("Score Fragmentacao",   "Score Fragmentação"),
        ("Score Lab",            "Score Lab"),
        ("Isotope Similarity",   "Score Isótopo"),
        ("Mass Error (ppm)",     "Erro de Massa (ppm)"),
        ("Score Massa",          "Score Massa (0–40)"),
        ("Adducts",              "Adducts"),
        ("Neutral Mass (Da)",    "Neutro Mass (Da)"),
        ("n_candidatos",         "N° de Candidatos"),
        ("Empate",               "Empate (S/N)"),
        ("Criterio Desempate",   "Critério de Desempate"),
        ("Score Qualidade Dados","Score Qualidade Dados"),
        ("Batch ID",             "Batch ID"),
    ]
    avail_src = [s for s, _ in desired if s in r1.columns]
    avail_dst = [d for s, d in desired if s in r1.columns]
    return r1[avail_src].rename(columns=dict(zip(avail_src, avail_dst)))


def _build_resultado_final(wb, ins: dict, nar: "NarrativaDict") -> None:
    ws = wb.create_sheet("Resultado_Final")
    df = _prep_resultado_df(ins, nar)
    _write_table(ws, df, priority_col="Prioridade")
    ws.freeze_panes = "A2"
    _auto_width(ws)


# ─── Aba 3: Para_Revisao ─────────────────────────────────────────────────

def _build_para_revisao(wb, ins: dict, nar: "NarrativaDict") -> None:
    ws = wb.create_sheet("Para_Revisao")

    prio_df = nar["priority_df"].copy()
    prio_df["Prioridade"] = prio_df["Prioridade"].map(_PRIO_LABEL).fillna("—")

    # n_candidatos do compound_data
    cd = ins["compound_data"][["Sinal", "n_candidatos"]].rename(
        columns={"Sinal": "Composto"}
    )
    prio_df = prio_df.merge(cd, on="Composto", how="left")
    prio_df.rename(columns={"n_candidatos": "N° Candidatos"}, inplace=True)

    # m/z do rank1_unico se disponível
    if "m/z Medido" in ins["rank1_unico"].columns:
        mz_map = (
            ins["rank1_unico"][["Sinal", "m/z Medido"]]
            .rename(columns={"Sinal": "Composto"})
        )
        prio_df = prio_df.merge(mz_map, on="Composto", how="left")

    revisao = prio_df[prio_df["Prioridade"].isin(["Alta", "Média"])].copy()

    if revisao.empty:
        ws.cell(row=1, column=1,
                value="Nenhum composto requer revisão. Análise aprovada.").font = (
            _font(bold=True, color="1b4f2a", size=11)
        )
        return

    revisao["Ação Necessária"] = revisao["Prioridade"].map(_ACAO)

    # Ordenação: Alta → Média, depois Confiança asc
    revisao["_ps"] = revisao["Prioridade"].map(_PRIO_SORT).fillna(3)
    revisao = revisao.sort_values(["_ps", "Confiança"], ascending=[True, True]).drop("_ps", axis=1)

    desired = ["Prioridade", "Composto"]
    if "m/z Medido" in revisao.columns:
        desired.append("m/z Medido")
    desired += ["Candidato mais provável", "Confiança", "N° Candidatos"]
    if "Categoria" in revisao.columns:
        desired.append("Categoria")
    desired += ["Situação", "Ação Necessária"]

    avail = [c for c in desired if c in revisao.columns]
    revisao = revisao[avail]

    _write_table(ws, revisao, priority_col="Prioridade")
    ws.freeze_panes = "A2"

    # Largura extra para "Ação Necessária"
    _auto_width(ws)
    if "Ação Necessária" in revisao.columns:
        idx_acao = revisao.columns.tolist().index("Ação Necessária") + 1
        ws.column_dimensions[get_column_letter(idx_acao)].width = 65


# ─── Aba 4: Dados_Tecnicos ───────────────────────────────────────────────

def _prep_dados_df(ins: dict) -> pd.DataFrame:
    df = ins["df"].copy()
    score_col = ins["score_col"]

    # Remove alias redundante se Score Ranking disponível
    if score_col == "Score Ranking" and "Score Total" in df.columns:
        df = df.drop("Score Total", axis=1, errors="ignore")

    # Ordena: Sinal asc → Rank asc → score desc
    sort_keys = [c for c in ["Sinal", "Rank", score_col] if c in df.columns]
    asc_flags = [True, True, False][:len(sort_keys)]
    if sort_keys:
        df = df.sort_values(sort_keys, ascending=asc_flags)

    rename = {
        "Sinal":                "Sinal",
        "m/z Medido":           "m/z",
        "Rank":                 "Rank",
        "Empate":               "Empate",
        score_col:              "Confiança (0–100)",
        "Score Fragmentacao":   "Score Fragmentação",
        "Score Lab":            "Score Lab",
        "Isotope Similarity":   "Score Isótopo",
        "Score Massa":          "Score Massa (0–40)",
        "Mass Error (ppm)":     "Erro de Massa (ppm)",
        "Score Qualidade Dados":"Score Qualidade Dados",
        "Candidato":            "Candidato",
        "Formula":              "Fórmula",
        "Peso Teorico":         "Peso Molecular (Da)",
        "Neutral Mass (Da)":    "Neutro Mass (Da)",
        "Adducts":              "Adducts",
        "Categoria":            "Categoria",
        "Classe Quimica":       "Classe Química (ChEBI)",
        "Criterio Desempate":   "Critério de Desempate",
        "Metodo Ranking":       "Método de Ranking",
        "Rank Group":           "Rank Group",
        "Batch ID":             "Batch ID",
        "Data Execucao":        "Data Execução",
    }
    avail = [c for c in rename if c in df.columns]
    return df[avail].rename(columns={c: rename[c] for c in avail})


def _build_dados_tecnicos(wb, ins: dict) -> None:
    ws = wb.create_sheet("Dados_Tecnicos")
    df = _prep_dados_df(ins)
    _write_table(ws, df)
    # Congela: linha 1 (header) + colunas A–C (Sinal, m/z, Rank)
    ws.freeze_panes = "D2"
    _auto_width(ws)


# ─── Aba 5: Estatisticas ─────────────────────────────────────────────────

def _write_mini_table(ws, df: pd.DataFrame, start_row: int,
                      prio_col: str | None = None) -> int:
    """Escreve tabela compacta sem auto-filtro. Retorna próxima linha livre."""
    cols = list(df.columns)
    _write_header(ws, cols, start_row)
    for r_off, (_, series) in enumerate(df.iterrows()):
        r = start_row + 1 + r_off
        row_fill = _C["alt"] if r_off % 2 == 1 else _C["white"]
        if prio_col and prio_col in cols:
            pval = str(series[prio_col])
            row_fill = _C.get(_PRIO_COLOR.get(pval, "white"), _C["white"])
        for c_off, val in enumerate(series):
            cell = ws.cell(row=r, column=c_off + 1, value=_norm(val))
            cell.alignment = _align()
            cell.fill = _fill(row_fill)
    return start_row + 1 + len(df)


def _build_estatisticas(wb, ins: dict, nar: "NarrativaDict",
                        cobertura_ext: dict) -> None:
    ws = wb.create_sheet("Estatisticas")
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 14

    n = 1
    n_comp    = ins["n_compostos"]
    score_col = ins["score_col"]

    # ── A: Distribuição por Prioridade ────────────────────────────────────
    n = _sect(ws, n, "DISTRIBUIÇÃO POR PRIORIDADE", ncols=3)
    prio_series = nar["priority_df"]["Prioridade"].map(_PRIO_LABEL).fillna("—")
    prio_counts = prio_series.value_counts().reset_index()
    prio_counts.columns = ["Prioridade", "N° Compostos"]
    prio_counts["% do Total"] = (
        prio_counts["N° Compostos"] / len(prio_series) * 100
    ).round(1).astype(str) + "%"
    prio_counts["_s"] = prio_counts["Prioridade"].map(_PRIO_SORT).fillna(3)
    prio_counts = prio_counts.sort_values("_s").drop("_s", axis=1)
    n = _write_mini_table(ws, prio_counts, n, prio_col="Prioridade")
    n += 1

    # ── B: Distribuição por Categoria Química ─────────────────────────────
    n = _sect(ws, n, "DISTRIBUIÇÃO POR CATEGORIA QUÍMICA", ncols=3)
    if not ins["classes_classif"].empty:
        cats = ins["classes_classif"].head(10).copy()
        cats.columns = ["Categoria", "N° Compostos"]
        cats["% do Total"] = (
            cats["N° Compostos"] / n_comp * 100
        ).round(1).astype(str) + "%"
        if ins["n_nc"] > 0:
            pct_nc = round(ins["n_nc"] / n_comp * 100, 1)
            cats.loc[len(cats)] = ["Não classificado", ins["n_nc"], f"{pct_nc}%"]
        n = _write_mini_table(ws, cats, n)
    else:
        ws.cell(row=n, column=1, value="Dados de classificação não disponíveis.")
        n += 1
    n += 1

    # ── C: Faixas de Confiança ─────────────────────────────────────────────
    n = _sect(ws, n, "FAIXAS DE CONFIANÇA (RANK 1)", ncols=3)
    scores = pd.to_numeric(
        ins["rank1_unico"][score_col], errors="coerce"
    ).dropna()
    total_sc = len(scores)
    alta  = int((scores >= 80).sum())
    media = int(((scores >= 45) & (scores < 80)).sum())
    baixa = int((scores < 45).sum())
    conf_data = pd.DataFrame([
        ["Alta  (≥ 80)",     alta,  f"{alta  / total_sc * 100:.0f}%" if total_sc else "0%"],
        ["Moderada (45–79)", media, f"{media / total_sc * 100:.0f}%" if total_sc else "0%"],
        ["Baixa  (< 45)",    baixa, f"{baixa / total_sc * 100:.0f}%" if total_sc else "0%"],
    ], columns=["Faixa de Confiança", "N° Compostos", "% do Total"])
    n = _write_mini_table(ws, conf_data, n)
    n += 1

    # ── D: Critérios de Desempate ─────────────────────────────────────────
    if not ins["criterio_counts"].empty:
        n = _sect(ws, n, "CRITÉRIOS DE DESEMPATE", ncols=3)
        crit = ins["criterio_counts"][["label", "n"]].copy()
        crit.columns = ["Critério", "N° Compostos"]
        crit["% do Total"] = (
            crit["N° Compostos"] / n_comp * 100
        ).round(1).astype(str) + "%"
        n = _write_mini_table(ws, crit, n)
        n += 1

    # ── E: Cobertura de Banco de Dados ───────────────────────────────────
    if cobertura_ext:
        n = _sect(ws, n, "COBERTURA DE BANCO DE DADOS", ncols=3)
        cov_rows = []
        if cobertura_ext.get("pct_pubchem", 0) > 0:
            cov_rows.append([
                "PubChem",
                cobertura_ext.get("com_pubchem", "—"),
                f"{cobertura_ext['pct_pubchem']:.0f}%",
            ])
        if cobertura_ext.get("pct_chebi", 0) > 0:
            cov_rows.append([
                "ChEBI",
                cobertura_ext.get("com_chebi", "—"),
                f"{cobertura_ext['pct_chebi']:.0f}%",
            ])
        if cov_rows:
            cov_df = pd.DataFrame(cov_rows,
                                  columns=["Base de Dados",
                                           "Compostos com dados", "% do Total"])
            n = _write_mini_table(ws, cov_df, n)

    # ── Largura automática (amostra) ─────────────────────────────────────
    _auto_width(ws, max_w=45)


# ─── Aba 6: Metadados ────────────────────────────────────────────────────

def _build_metadados(wb, batch_info: dict, cobertura_ext: dict,
                     ins: dict) -> None:
    ws = wb.create_sheet("Metadados")
    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 60

    n = 1

    # Título
    t = ws.cell(row=n, column=1, value="Metadados do Experimento e Rastreabilidade")
    t.font  = _font(bold=True, color=_C["hdr_fg"], size=12)
    t.fill  = _fill(_C["hdr_bg"])
    t.alignment = _align(h="left")
    ws.merge_cells(start_row=n, start_column=1, end_row=n, end_column=2)
    ws.row_dimensions[n].height = 22
    n += 2

    n = _sect(ws, n, "IDENTIFICAÇÃO DO BATCH", ncols=2)
    d_ini = (batch_info.get("iniciado_em") or "")[:16].replace("T", " ") or "—"
    d_fim = (batch_info.get("concluido_em") or "")[:16].replace("T", " ") or "—"
    n = _kv(ws, n, "Análise ID",              batch_info.get("id", "—"))
    n = _kv(ws, n, "Status do processamento", batch_info.get("status", "—"))
    n = _kv(ws, n, "Iniciado em",             d_ini)
    n = _kv(ws, n, "Concluído em",            d_fim)
    n = _kv(ws, n, "Fonte",                   batch_info.get("fonte", "—"))
    n += 1

    n = _sect(ws, n, "ARQUIVOS DE ENTRADA", ncols=2)
    n = _kv(ws, n, "Arquivo de identificação", batch_info.get("nome_ident", "—"))
    n = _kv(ws, n, "Arquivo de abundância",    batch_info.get("nome_abund", "—"))
    n = _kv(ws, n, "SHA-256 Identificação",    batch_info.get("hash_ident", "—"))
    n = _kv(ws, n, "SHA-256 Abundância",       batch_info.get("hash_abund", "—"))
    n += 1

    n = _sect(ws, n, "ESTATÍSTICAS DO PROCESSAMENTO", ncols=2)
    n = _kv(ws, n, "Total de sinais processados",
             batch_info.get("total_sinais") or ins["n_compostos"])
    n = _kv(ws, n, "Total de candidatos avaliados",
             batch_info.get("total_candidatos") or ins["n_candidatos_tot"])
    n = _kv(ws, n, "Moléculas consultadas via API",
             batch_info.get("total_moleculas_api", "—"))
    if cobertura_ext:
        n = _kv(ws, n, "Cobertura ChEBI",
                 f"{cobertura_ext.get('pct_chebi', 0):.0f}%")
        n = _kv(ws, n, "Cobertura PubChem",
                 f"{cobertura_ext.get('pct_pubchem', 0):.0f}%")
    n += 1

    n = _sect(ws, n, "MÉTODO DE ANÁLISE", ncols=2)
    n = _kv(ws, n, "Algoritmo de ranking",    "Hierárquico IST v4")
    n = _kv(ws, n, "Pesos (frag/lab/iso/massa)", "40% / 30% / 20% / 10%")
    n = _kv(ws, n, "Limiar Alta confiança",   "Score ≥ 80")
    n = _kv(ws, n, "Limiar Baixa confiança",  "Score < 45")
    n += 1

    n = _sect(ws, n, "EXPORTAÇÃO", ncols=2)
    n = _kv(ws, n, "Exportado em",
             datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
    n = _kv(ws, n, "Software", "Omics ETL Pipeline · IST Ambiental / SENAI")
    _kv(ws, n, "Aviso",
        "Resultados para uso interno. Requerem validação do especialista analítico.")


# ─── Ponto de entrada público ─────────────────────────────────────────────

def gerar_exportacao_xlsx(
    ins: dict,
    nar: "NarrativaDict",
    batch_info: dict | None = None,
    cobertura_ext: dict | None = None,
) -> bytes:
    """
    Gera workbook XLSX consolidado com 6 abas.
    Retorna bytes prontos para st.download_button ou escrita em arquivo.
    """
    if not _HAS_OPENPYXL:
        raise ImportError(
            "openpyxl não encontrado. Instale com: pip install openpyxl"
        )

    batch_info    = batch_info    or {}
    cobertura_ext = cobertura_ext or {}

    wb = Workbook()
    wb.remove(wb.active)   # remove aba padrão

    _build_resumo(wb, ins, nar, batch_info, cobertura_ext)
    _build_resultado_final(wb, ins, nar)
    _build_para_revisao(wb, ins, nar)
    _build_dados_tecnicos(wb, ins)
    _build_estatisticas(wb, ins, nar, cobertura_ext)
    _build_metadados(wb, batch_info, cobertura_ext, ins)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
